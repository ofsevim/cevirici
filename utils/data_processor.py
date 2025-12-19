"""
Veri İşleme Yardımcı Fonksiyonlar
Bu modül, dosya okuma, karakter düzeltme ve veri temizleme işlemlerini içerir.
"""

import pandas as pd
import re


def fix_turkish_chars(text):
    """
    Bozuk encoding'den kaynaklı Türkçe karakter hatalarını düzeltir.
    
    Args:
        text (str): Düzeltilecek metin
    
    Returns:
        str: Düzeltilmiş metin
    """
    if not isinstance(text, str):
        return text
    
    # Yaygın encoding hataları haritası
    replacements = {
        'Ã¼': 'ü', 'Ã¶': 'ö', 'Ã§': 'ç', 'ÅŸ': 'ş', 'Ä±': 'ı', 'ÄŸ': 'ğ',
        'Ãœ': 'Ü', 'Ã–': 'Ö', 'Ã‡': 'Ç', 'Åž': 'Ş', 'Ä°': 'İ', 'Äž': 'Ğ',
        'Ý': 'İ', 'Þ': 'Ş', 'ð': 'ğ', 'ý': 'ı', 'þ': 'ş', 'Ð': 'Ğ'
    }
    
    for bad, good in replacements.items():
        text = text.replace(bad, good)
    
    return text


def read_file_with_encoding(uploaded_file, skip_rows=0):
    """
    Yüklenen dosyayı uygun encoding ile okur.
    Excel ve metin dosyalarını destekler.
    
    Args:
        uploaded_file: Streamlit file uploader objesi
        skip_rows (int): Atlanacak başlangıç satır sayısı
    
    Returns:
        pd.DataFrame: Ham veri DataFrame'i
    """
    
    # Excel dosyası kontrolü
    if uploaded_file.name.endswith('.xlsx') or uploaded_file.name.endswith('.xls'):
        try:
            # İlk olarak openpyxl ile merged cell bilgisini alalım (sadece .xlsx için)
            if uploaded_file.name.endswith('.xlsx'):
                try:
                    from openpyxl import load_workbook
                    from io import BytesIO
                    
                    # Dosyayı oku
                    uploaded_file.seek(0)
                    wb = load_workbook(BytesIO(uploaded_file.read()), data_only=True)
                    ws = wb.active
                    
                    # Merged cell aralıklarını kaydet
                    merged_ranges = list(ws.merged_cells.ranges)
                    
                    # Merged cell'leri unmerge et ve değerleri yay
                    for merged_range in merged_ranges:
                        # Merged aralığın ilk hücresindeki değeri al
                        top_left_cell = ws.cell(merged_range.min_row, merged_range.min_col)
                        top_left_value = top_left_cell.value
                        
                        # Önce unmerge et
                        ws.unmerge_cells(str(merged_range))
                        
                        # Tüm hücrelere aynı değeri yaz
                        for row in range(merged_range.min_row, merged_range.max_row + 1):
                            for col in range(merged_range.min_col, merged_range.max_col + 1):
                                ws.cell(row, col, top_left_value)
                    
                    # DataFrame'e dönüştür
                    data = []
                    for row in ws.iter_rows(min_row=skip_rows + 1, values_only=True):
                        data.append(list(row))
                    
                    df = pd.DataFrame(data, dtype=str)
                    
                    wb.close()
                    
                except Exception as e:
                    # openpyxl başarısız olursa normal pandas ile oku
                    uploaded_file.seek(0)
                    df = pd.read_excel(uploaded_file, header=None, dtype=str, skiprows=skip_rows)
            else:
                # .xls dosyaları için normal okuma
                df = pd.read_excel(uploaded_file, header=None, dtype=str, skiprows=skip_rows)
            
            # None değerlerini NaN'a çevir (openpyxl'den gelen)
            df = df.replace(['None', 'none', ''], pd.NA)
            
            # Tamamen boş satırları temizle
            df = df.dropna(how='all').reset_index(drop=True)
            
            # Tamamen boş sütunları temizle
            df = df.dropna(axis=1, how='all')
            
            # Sütun numaralarını yeniden düzenle
            df.columns = range(len(df.columns))
            
            return df
        except Exception as e:
            raise ValueError(f"Excel okuma hatası: {e}")
    
    # Metin dosyası (CSV/TXT) için encoding denemeleri
    raw_bytes = uploaded_file.getvalue()
    
    encodings = ['cp1254', 'utf-8', 'iso-8859-9', 'latin-1']
    
    for enc in encodings:
        try:
            string_data = raw_bytes.decode(enc)
            
            # Atlanacak satırları çıkar
            if skip_rows > 0:
                lines = string_data.split('\n')
                string_data = '\n'.join(lines[skip_rows:])
            
            # Ayırıcıyı tespit et
            if ';' in string_data.split('\n')[0]:
                separator = ';'
            elif ',' in string_data.split('\n')[0]:
                separator = ','
            elif '\t' in string_data.split('\n')[0]:
                separator = '\t'
            else:
                separator = ','
            
            # DataFrame'e dönüştür
            from io import StringIO
            df = pd.read_csv(StringIO(string_data), sep=separator, header=None, dtype=str, engine='python')
            
            # Tamamen boş satırları temizle
            df = df.dropna(how='all').reset_index(drop=True)
            # Tamamen boş sütunları temizle
            df = df.dropna(axis=1, how='all')
            # Sütun numaralarını yeniden düzenle
            df.columns = range(len(df.columns))
            
            return df
            
        except UnicodeDecodeError:
            continue
        except Exception:
            continue
    
    raise ValueError("Dosya okunamadı. Desteklenen formatlar: CSV, TXT, XLSX, XLS")


def clean_amount_value(value):
    """
    Tutar değerini temizler ve sayıya çevirir.
    
    Args:
        value (str): Ham tutar değeri
    
    Returns:
        float: Temizlenmiş tutar
    """
    if pd.isna(value) or value == "" or value is None:
        return 0.0
    
    value_str = str(value).strip()
    
    # Boş string kontrolü
    if not value_str or value_str in ['None', 'nan', 'NaN', 'null', '-']:
        return 0.0
    
    # Tırnak işaretlerini temizle
    value_str = value_str.replace('"', '').replace("'", "").strip()
    
    # Para birimi sembollerini temizle
    value_str = value_str.replace('₺', '').replace('TL', '').replace('TRY', '').strip()
    
    # Binlik ayracı ve ondalık ayracı kontrolü
    # Türkçe format: 1.234,56 -> İngilizce format: 1234.56
    # İngilizce format: 1,234.56 -> 1234.56
    
    has_comma = ',' in value_str
    has_dot = '.' in value_str
    
    if has_comma and has_dot:
        # Her iki ayraç da var - hangisi ondalık?
        comma_pos = value_str.rfind(',')
        dot_pos = value_str.rfind('.')
        
        if comma_pos > dot_pos:
            # Türkçe format: 1.234,56 -> nokta binlik, virgül ondalık
            value_str = value_str.replace('.', '')  # Binlik ayracını kaldır
            value_str = value_str.replace(',', '.')  # Ondalık ayracını noktaya çevir
        else:
            # İngilizce format: 1,234.56 -> virgül binlik, nokta ondalık
            value_str = value_str.replace(',', '')  # Binlik ayracını kaldır
    elif has_comma:
        # Sadece virgül var - muhtemelen Türkçe ondalık ayracı
        value_str = value_str.replace(',', '.')
    # Sadece nokta varsa zaten doğru format
    
    # Sadece sayı, nokta ve eksi işareti bırak
    value_str = re.sub(r'[^\d.\-]', '', value_str)
    
    # Birden fazla nokta varsa (örn: 1.234.567) son noktayı ondalık olarak kabul et
    if value_str.count('.') > 1:
        parts = value_str.rsplit('.', 1)
        value_str = parts[0].replace('.', '') + '.' + parts[1]
    
    # Boş string kontrolü
    if not value_str or value_str == '.' or value_str == '-':
        return 0.0
    
    try:
        return float(value_str)
    except ValueError:
        return 0.0


def split_full_name(full_name):
    """
    Tam adı (Ad Soyad) ayrı ayrı ad ve soyad olarak ayırır.
    
    Args:
        full_name (str): Tam ad (örn: "Ahmet Yılmaz" veya "Nazire Asil")
    
    Returns:
        tuple: (first_name, last_name)
    """
    if not full_name or full_name in ['None', 'nan', 'NaN', '']:
        return "", ""
    
    full_name = str(full_name).strip()
    
    # Boşluğa göre ayır
    parts = full_name.split()
    
    if len(parts) == 0:
        return "", ""
    elif len(parts) == 1:
        return parts[0], ""
    elif len(parts) == 2:
        return parts[0], parts[1]
    else:
        # 2'den fazla kelime varsa, son kelime soyad, geri kalanı ad
        first_name = " ".join(parts[:-1])
        last_name = parts[-1]
        return first_name, last_name


def clean_tc_number(value):
    """
    TC Kimlik numarasını temizler ve doğrular.
    
    Args:
        value (str): Ham TC değeri
    
    Returns:
        str: Temizlenmiş 11 haneli TC (geçerli değilse boş string)
    """
    if pd.isna(value):
        return ""
    
    value_str = str(value).strip()
    
    # Sadece rakamları al
    digits = re.sub(r'\D', '', value_str)
    
    # 11 hane kontrolü
    if len(digits) == 11:
        return digits
    
    return ""


def apply_column_mapping(df_raw, column_mapping):
    """
    Kullanıcının yaptığı sütun eşleştirmesine göre veriyi işler.
    
    Args:
        df_raw (pd.DataFrame): Ham veri
        column_mapping (dict): Sütun eşleştirme haritası
    
    Returns:
        tuple: (pd.DataFrame: Temizlenmiş veri, dict: İşlem istatistikleri)
    """
    
    processed_data = []
    stats = {
        'total_rows': len(df_raw),
        'processed_rows': 0,
        'skipped_rows': 0,
        'invalid_tc': 0,
        'empty_rows': 0,
        'zero_amount': 0,
        'sample_skipped': [],
        'sample_amounts': []  # Debug için örnek tutar bilgileri
    }
    
    use_combined_name = column_mapping.get('use_combined_name', False)
    
    for idx, row in df_raw.iterrows():
        try:
            # Tamamen boş satır kontrolü
            if row.isna().all():
                stats['empty_rows'] += 1
                continue
            
            # Her alan için mapping'e göre veriyi al
            member_no = str(row[column_mapping.get('member_no', 0)]).strip() if 'member_no' in column_mapping else ""
            tc_no = str(row[column_mapping.get('tc_no', 0)]).strip() if 'tc_no' in column_mapping else ""
            amount = str(row[column_mapping.get('amount', 0)]).strip() if 'amount' in column_mapping else "0"
            
            # Ad-Soyad işleme
            if use_combined_name and 'full_name' in column_mapping:
                # Birleşik ad/soyad varsa ayır
                full_name = str(row[column_mapping.get('full_name', 0)]).strip()
                first_name, last_name = split_full_name(full_name)
            else:
                # Ayrı sütunlar varsa direkt al
                first_name = str(row[column_mapping.get('first_name', 0)]).strip() if 'first_name' in column_mapping else ""
                last_name = str(row[column_mapping.get('last_name', 0)]).strip() if 'last_name' in column_mapping else ""
            
            # None veya nan kontrolü
            if member_no in ['None', 'nan', 'NaN']:
                member_no = ""
            if first_name in ['None', 'nan', 'NaN']:
                first_name = ""
            if last_name in ['None', 'nan', 'NaN']:
                last_name = ""
            if tc_no in ['None', 'nan', 'NaN']:
                tc_no = ""
            if amount in ['None', 'nan', 'NaN']:
                amount = "0"
            
            # Türkçe karakter düzeltmeleri
            first_name = fix_turkish_chars(first_name)
            last_name = fix_turkish_chars(last_name)
            
            # TC temizle
            tc_no_original = tc_no
            tc_no = clean_tc_number(tc_no)
            
            # Tutar temizle
            amount_original = amount
            amount_clean = clean_amount_value(amount)
            
            # Debug için örnek tutar bilgisi kaydet
            if tc_no:
                # Boş olmayan tutar örnekleri (ilk 5)
                if len(stats['sample_amounts']) < 5 and amount_clean > 0:
                    stats['sample_amounts'].append({
                        'satir': idx + 1,
                        'ham_tutar': amount_original,
                        'temiz_tutar': amount_clean
                    })
                
                # Sıfıra dönüşen tutar örnekleri (ilk 3) - debug için
                if 'sample_zero_amounts' not in stats:
                    stats['sample_zero_amounts'] = []
                if len(stats['sample_zero_amounts']) < 3 and amount_clean == 0.0:
                    stats['sample_zero_amounts'].append({
                        'satir': idx + 1,
                        'ham_tutar': amount_original
                    })
            
            # Sıfır tutar sayacı
            if amount_clean == 0.0 and amount_original not in ['', '0', 'nan', 'NaN', 'None']:
                stats['zero_amount'] += 1
            
            # Geçerli TC varsa ekle
            if tc_no and len(tc_no) == 11:
                processed_data.append({
                    "Üye No": member_no,
                    "Adı": first_name,
                    "Soyadı": last_name,
                    "TC Kimlik No": tc_no,
                    "Aidat Tutarı": amount_clean
                })
                stats['processed_rows'] += 1
            else:
                stats['invalid_tc'] += 1
                if len(stats['sample_skipped']) < 5:
                    stats['sample_skipped'].append({
                        'satir': idx + 1,
                        'tc': tc_no_original,
                        'ad': first_name,
                        'soyad': last_name
                    })
        
        except Exception as e:
            stats['skipped_rows'] += 1
            continue
    
    # DataFrame oluştur
    df_clean = pd.DataFrame(processed_data)
    
    return df_clean, stats


def find_data_start_row(uploaded_file, max_rows_to_check=50):
    """
    Excel/CSV dosyasında gerçek verinin başladığı satırı bulur.
    
    Args:
        uploaded_file: Streamlit file uploader objesi
        max_rows_to_check (int): Kontrol edilecek maksimum satır sayısı
    
    Returns:
        int: Veri başlangıç satırı (0-indexed)
    """
    try:
        # Dosyayı header olmadan oku
        if uploaded_file.name.endswith('.xlsx') or uploaded_file.name.endswith('.xls'):
            df_temp = pd.read_excel(uploaded_file, header=None, nrows=max_rows_to_check, dtype=str)
        else:
            uploaded_file.seek(0)
            raw_bytes = uploaded_file.getvalue()
            string_data = raw_bytes.decode('cp1254', errors='ignore')
            
            from io import StringIO
            df_temp = pd.read_csv(StringIO(string_data), header=None, nrows=max_rows_to_check, dtype=str, sep=None, engine='python')
        
        # Her satırı analiz et
        for idx, row in df_temp.iterrows():
            # Satırdaki dolu hücre sayısı
            non_empty = row.notna().sum()
            
            # En az 3 sütun dolu ise ve TC kimlik pattern'i varsa
            if non_empty >= 3:
                row_str = ' '.join(row.dropna().astype(str).tolist())
                
                # TC Kimlik var mı kontrol et (11 haneli sayı)
                if re.search(r'\d{11}', row_str):
                    return idx
                
                # Veya çok sayıda sayısal veri varsa (tablo başlangıcı olabilir)
                numeric_count = sum(1 for val in row.dropna() if str(val).replace(',', '').replace('.', '').isdigit())
                if numeric_count >= 2:
                    return idx
        
        return 0
        
    except Exception:
        return 0


def detect_file_structure(df_raw, sample_size=50):
    """
    Ham veriyi analiz eder ve sütun yapısı hakkında bilgi verir.
    
    Args:
        df_raw (pd.DataFrame): Ham veri
        sample_size (int): Analiz edilecek satır sayısı
    
    Returns:
        dict: Dosya yapısı bilgileri
    """
    
    info = {
        'total_rows': len(df_raw),
        'total_columns': len(df_raw.columns),
        'has_tc_column': False,
        'tc_column_index': None,
        'has_amount_column': False,
        'sample_data': df_raw.head(sample_size)
    }
    
    # TC Kimlik içeren sütun var mı?
    for col_idx, col in enumerate(df_raw.columns):
        sample_values = df_raw[col].astype(str).head(sample_size)
        
        # TC pattern (11 hane)
        tc_count = sample_values.str.match(r'^\d{11}$').sum()
        
        if tc_count > sample_size * 0.5:  # %50'den fazlası TC ise
            info['has_tc_column'] = True
            info['tc_column_index'] = col_idx
            break
    
    # Tutar içeren sütun var mı?
    for col in df_raw.columns:
        sample_values = df_raw[col].astype(str).head(sample_size)
        
        # Sayısal pattern
        numeric_count = sample_values.str.match(r'^[\d\.,]+$').sum()
        
        if numeric_count > sample_size * 0.5:
            info['has_amount_column'] = True
            break
    
    return info

